from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = REPO_ROOT / "docs" / "validation-data" / "validation-workbook.xlsx"
OUTPUT_WORKBOOK = REPO_ROOT / "docs" / "validation-data" / "validation-workbook-small.xlsx"
TARGET_ORDER_COUNT = 42
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


def load_order_entries(order_ws):
    headers = [order_ws.cell(row=1, column=col).value for col in range(1, order_ws.max_column + 1)]
    entries = []
    for row_index in range(2, order_ws.max_row + 1):
        values = [order_ws.cell(row=row_index, column=col).value for col in range(1, order_ws.max_column + 1)]
        entries.append({
            "row_index": row_index,
            "values": values,
        })
    return headers, entries


def contains_tag(entry, tag: str) -> bool:
    value = entry["values"][17] or ""
    return tag in str(value)


def get_line(entry) -> str:
    return str(entry["values"][8] or "")


def get_product(entry) -> str:
    return str(entry["values"][11] or "")


def get_thickness(entry) -> int:
    try:
        return int(entry["values"][10])
    except (TypeError, ValueError):
        return -1


def get_duration(entry) -> float:
    try:
        return float(entry["values"][12])
    except (TypeError, ValueError):
        return 0.0


def get_compatible_lines(entry) -> str:
    return str(entry["values"][15] or "")


def get_source_tag(entry) -> str:
    return str(entry["values"][20] or "")


def is_synthetic(entry) -> bool:
    return get_source_tag(entry).startswith("synthetic-")


def is_dual_compatible(entry) -> bool:
    return ";" in get_compatible_lines(entry)


def select_entries(entries):
    selected = []
    selected_ids = set()

    def add(entry):
        order_id = str(entry["values"][0] or "")
        if not order_id or order_id in selected_ids:
            return False
        selected.append(entry)
        selected_ids.add(order_id)
        return True

    synthetic_entries = [entry for entry in entries if is_synthetic(entry)]
    base_entries = [entry for entry in entries if not is_synthetic(entry)]

    for entry in synthetic_entries:
        add(entry)

    for product_code in ["DJX", "DJY", "EST", "ESY", "FDX", "FDY", "QDJY"]:
        for entry in base_entries:
            if get_product(entry) == product_code:
                add(entry)
                break

    for line_code in ["双拉2线", "双拉4线"]:
        for tag in ["HC2-urgent", "MC2-high-stock"]:
            for entry in base_entries:
                if get_line(entry) == line_code and contains_tag(entry, tag):
                    add(entry)
                    break

    for product_code in ["EST", "ESY", "FDX", "FDY", "DJX", "DJY"]:
        for entry in base_entries:
            if get_product(entry) == product_code and is_dual_compatible(entry):
                add(entry)
                break

    thick_targets = [4, 5, 7, 9, 10, 14, 24, 29, 42, 61]
    for thickness in thick_targets:
        for entry in base_entries:
            if get_thickness(entry) == thickness:
                add(entry)
                break

    long_base_entries = sorted(
        [entry for entry in base_entries if get_duration(entry) >= 48.0],
        key=lambda item: (-get_duration(item), str(item["values"][0] or ""))
    )
    for entry in long_base_entries:
        if len(selected) >= TARGET_ORDER_COUNT:
            break
        add(entry)

    prioritized_fill = sorted(
        base_entries,
        key=lambda item: (
            0 if is_dual_compatible(item) else 1,
            0 if contains_tag(item, "HC2-urgent") else 1,
            0 if contains_tag(item, "MC2-high-stock") else 1,
            0 if get_line(item) == "双拉2线" else 1,
            str(item["values"][0] or "")
        )
    )
    for entry in prioritized_fill:
        if len(selected) >= TARGET_ORDER_COUNT:
            break
        add(entry)

    return selected


def copy_header_row(source_ws, target_ws):
    for col in range(1, source_ws.max_column + 1):
        source_cell = source_ws.cell(row=1, column=col)
        target_cell = target_ws.cell(row=1, column=col, value=source_cell.value)
        target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.border = copy(source_cell.border)
    target_ws.freeze_panes = source_ws.freeze_panes

    for column_key, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[column_key].width = dimension.width


def copy_selected_orders(source_ws, target_wb, selected_entries):
    target_ws = target_wb.create_sheet(source_ws.title)
    copy_header_row(source_ws, target_ws)

    for target_row_index, entry in enumerate(selected_entries, start=2):
        source_row_index = entry["row_index"]
        for col in range(1, source_ws.max_column + 1):
            source_cell = source_ws.cell(row=source_row_index, column=col)
            target_cell = target_ws.cell(row=target_row_index, column=col, value=source_cell.value)
            target_cell._style = copy(source_cell._style)
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.border = copy(source_cell.border)
            target_cell.number_format = source_cell.number_format
    return target_ws


def copy_full_sheet(source_ws, target_wb):
    target_ws = target_wb.create_sheet(source_ws.title)
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws.cell(row=source_cell.row, column=source_cell.column, value=source_cell.value)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.border = copy(source_cell.border)

    target_ws.freeze_panes = source_ws.freeze_panes
    for column_key, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[column_key].width = dimension.width
    for row_key, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_key].height = dimension.height
    return target_ws


def style_simple_sheet(ws):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)


def build_summary_sheet(target_wb, selected_entries, source_wb):
    ws = target_wb.create_sheet("统计摘要")
    dual_count = sum(1 for entry in selected_entries if is_dual_compatible(entry))
    urgent_count = sum(1 for entry in selected_entries if contains_tag(entry, "HC2-urgent"))
    high_stock_count = sum(1 for entry in selected_entries if contains_tag(entry, "MC2-high-stock"))
    long_count = sum(1 for entry in selected_entries if get_duration(entry) > 72.0)
    synthetic_count = sum(1 for entry in selected_entries if is_synthetic(entry))

    ws.append(["metric", "value", "note"])
    ws.append(["selected_orders", len(selected_entries), "Targeted small regression dataset"])
    ws.append(["synthetic_orders", synthetic_count, "Includes MC1 and split candidates"])
    ws.append(["dual_compatible_orders", dual_count, "For HC1 flexible routing checks"])
    ws.append(["urgent_inventory_orders", urgent_count, "For HC2 or SC2 style verification"])
    ws.append(["high_inventory_orders", high_stock_count, "For MC2 style verification"])
    ws.append(["long_duration_orders", long_count, "For HC5/SC5 candidate analysis"])
    ws.append(["exception_windows", source_wb[source_wb.sheetnames[1]].max_row - 1, "Copied from full validation workbook"])
    ws.append(["filter_change_plans", source_wb[source_wb.sheetnames[2]].max_row - 1, "Copied from full validation workbook"])
    ws.append(["calendar_rows", source_wb[source_wb.sheetnames[3]].max_row - 1, "Copied from full validation workbook"])
    style_simple_sheet(ws)


def build_coverage_sheet(target_wb):
    ws = target_wb.create_sheet("约束覆盖矩阵")
    ws.append(["constraint", "status", "coverage"])
    rows = [
        ("HC1", "ready", "Contains single-line and dual-compatible orders"),
        ("HC2", "ready", "Contains urgent inventory rows"),
        ("HC3", "ready", "Includes exception time sheet"),
        ("HC4", "ready", "Contains multiple thickness transitions"),
        ("MC1", "ready", "Contains synthetic priority sequence rows"),
        ("MC2", "ready", "Contains high-inventory rows"),
        ("SC1", "ready", "Contains changeover-relevant product and thickness changes"),
        ("SC2", "ready", "Contains different stock-day levels"),
        ("SC3", "ready", "Contains varied start expectations"),
        ("SC4", "ready", "Contains preferred vs compatible line differences"),
        ("HC5", "candidate", "Contains long-duration orders for split-mode testing"),
        ("SC5", "candidate", "Contains long-duration orders for split continuity testing"),
    ]
    for row in rows:
        ws.append(row)
    style_simple_sheet(ws)


def main():
    if not SOURCE_WORKBOOK.exists():
        raise FileNotFoundError(f"Full validation workbook not found: {SOURCE_WORKBOOK}")

    source_wb = load_workbook(SOURCE_WORKBOOK, data_only=False)
    sheetnames = source_wb.sheetnames
    order_ws = source_wb[sheetnames[0]]
    exception_ws = source_wb[sheetnames[1]]
    filter_ws = source_wb[sheetnames[2]]
    calendar_ws = source_wb[sheetnames[3]]

    _, entries = load_order_entries(order_ws)
    selected_entries = select_entries(entries)

    target_wb = Workbook()
    target_wb.remove(target_wb.active)

    copy_selected_orders(order_ws, target_wb, selected_entries)
    copy_full_sheet(exception_ws, target_wb)
    copy_full_sheet(filter_ws, target_wb)
    copy_full_sheet(calendar_ws, target_wb)
    build_coverage_sheet(target_wb)
    build_summary_sheet(target_wb, selected_entries, source_wb)

    OUTPUT_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(OUTPUT_WORKBOOK)

    line_counts = {}
    for entry in selected_entries:
        line_counts[get_line(entry)] = line_counts.get(get_line(entry), 0) + 1

    print(f"Generated small validation workbook: {OUTPUT_WORKBOOK}")
    print(f"Selected orders: {len(selected_entries)}")
    print(f"Line distribution: {line_counts}")


if __name__ == "__main__":
    main()

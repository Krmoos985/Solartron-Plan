from __future__ import annotations

from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = REPO_ROOT / "src" / "main" / "resources" / "data" / "生产订单_排程导入版.xlsx"
OUTPUT_DIR = REPO_ROOT / "docs" / "validation-data"
OUTPUT_WORKBOOK = OUTPUT_DIR / "validation-workbook.xlsx"

BASE_START = datetime(2026, 3, 13, 8, 0)
DATE_FMT = "%Y.%m.%d"
DATETIME_FMT = "%Y-%m-%d %H:%M"

ORIGINAL_HEADERS = [
    "订单",
    "物料",
    "物料描述",
    "订单数量(M2)",
    "确认产量",
    "已交货",
    "排产开始",
    "计划完工",
    "线别",
    "配方编码",
    "厚度(μm)",
    "型号",
    "生产时长(小时)",
    "当前库存(M2)",
    "月发货量(M2)",
    "兼容产线",
    "备注",
]

EXTRA_HEADERS = [
    "验证标签",
    "目标库存覆盖天数",
    "兼容策略",
    "伪造来源",
]

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")

MC1_RANK_COMBOS = [
    ("EST", "Formula_EST", 19, "双拉4线", "MC1-rank-1"),
    ("FDX", "Formula_FD", 4, "双拉4线", "MC1-rank-2"),
    ("FDY", "Formula_FD", 4, "双拉4线", "MC1-rank-2"),
    ("FDX", "Formula_FD", 5, "双拉4线", "MC1-rank-3"),
    ("FDX", "Formula_FD", 7, "双拉4线", "MC1-rank-4"),
    ("DJX", "Formula_DJ", 24, "双拉4线", "MC1-rank-5"),
    ("EST", "Formula_EST", 9, "双拉4线", "MC1-rank-6"),
]


def pick_target_days(index_zero_based: int) -> int:
    cycle = [5, 8, 12, 18, 25, 40, 60]
    return cycle[index_zero_based % len(cycle)]


def pick_monthly_factor(product_code: str) -> float:
    if product_code in {"EST", "ESY"}:
        return 0.60
    if product_code in {"FDX", "FDY"}:
        return 0.55
    return 0.50


def build_inventory_fields(quantity: float, product_code: str, target_days: int) -> tuple[float, float]:
    monthly = round(max(quantity * pick_monthly_factor(product_code), 50_000.0), 2)
    inventory = round(monthly * target_days / 30.0, 2)
    return inventory, monthly


def build_compatibility(line_code: str, product_code: str, thickness: int, index_zero_based: int) -> tuple[str, str]:
    if product_code == "QDJY":
        return "双拉2线", "single-L2"

    if product_code in {"EST", "ESY"} and thickness in {9, 10, 14}:
        return "双拉2线;双拉4线", "dual-family-est"

    if product_code in {"FDX", "FDY"} and thickness in {4, 5, 7}:
        return "双拉2线;双拉4线", "dual-family-fd"

    if index_zero_based % 5 == 0:
        return "双拉2线;双拉4线", "dual-interval"

    return line_code, f"single-{line_code}"


def build_validation_tags(target_days: int, compat_strategy: str, duration_hours: float) -> str:
    tags: list[str] = []

    if target_days < 10:
        tags.append("HC2-urgent")
    elif target_days > 30:
        tags.append("MC2-high-stock")
    else:
        tags.append("stock-balanced")

    if compat_strategy.startswith("dual"):
        tags.append("HC1-flex")

    if duration_hours > 72:
        tags.append("HC5-SC5-split-candidate")

    return "|".join(tags)


def style_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 32)


def create_orders_sheet(source_ws, target_wb: Workbook) -> tuple[int, Counter[str]]:
    ws = target_wb.create_sheet("订单_验证版")
    ws.append(ORIGINAL_HEADERS + EXTRA_HEADERS)

    stats = Counter()
    synthetic_count = 0

    for row_index in range(2, source_ws.max_row + 1):
        values = [source_ws.cell(row=row_index, column=col).value for col in range(1, 18)]
        quantity = float(values[3])
        line_code = str(values[8])
        formula_code = str(values[9])
        thickness = int(values[10])
        product_code = str(values[11])
        duration_hours = float(values[12])

        target_days = pick_target_days(row_index - 2)
        inventory, monthly = build_inventory_fields(quantity, product_code, target_days)
        compat_value, compat_strategy = build_compatibility(line_code, product_code, thickness, row_index - 2)
        validation_tags = build_validation_tags(target_days, compat_strategy, duration_hours)

        values[13] = inventory
        values[14] = monthly
        values[15] = compat_value
        values[16] = f"BASE|INV={target_days}d|COMPAT={compat_strategy}"

        ws.append(values + [validation_tags, target_days, compat_strategy, "base-468"])

        stats["base_rows"] += 1
        if target_days < 10:
            stats["urgent_rows"] += 1
        if target_days > 30:
            stats["high_stock_rows"] += 1
        if compat_strategy.startswith("dual"):
            stats["dual_compatible_rows"] += 1
        else:
            stats["single_compatible_rows"] += 1

    for idx, (product_code, formula_code, thickness, line_code, scenario_tag) in enumerate(MC1_RANK_COMBOS, start=1):
        quantity = 220_000.0 + idx * 15_000.0
        duration = round(max(quantity / 4_000.0, 18.0), 1)
        target_days = 8 if idx <= 2 else 18
        inventory, monthly = build_inventory_fields(quantity, product_code, target_days)
        compat_value = "双拉2线;双拉4线"
        compat_strategy = "dual-synthetic"
        order_id = f"VAL-MC1-{idx:03d}"

        ws.append([
            order_id,
            f"VAL-MAT-{product_code}-{thickness}",
            f"验证膜_T{thickness}_{product_code}_{idx:03d}",
            quantity,
            round(quantity * 0.95, 2),
            0,
            BASE_START.strftime(DATE_FMT),
            BASE_START.strftime(DATE_FMT),
            line_code,
            formula_code,
            thickness,
            product_code,
            duration,
            inventory,
            monthly,
            compat_value,
            f"SYNTHETIC|{scenario_tag}",
            f"MC1|{scenario_tag}",
            target_days,
            compat_strategy,
            "synthetic-mc1",
        ])
        synthetic_count += 1

    extra_long_orders = [
        ("VAL-SPLIT-001", "DJY", "Formula_DJ", 29, "双拉2线", 2_400_000.0, 96.0, 5, "HC5-SC5-long-order"),
        ("VAL-SPLIT-002", "EST", "Formula_EST", 61, "双拉4线", 1_900_000.0, 84.0, 45, "HC5-SC5-long-order"),
    ]

    for order_id, product_code, formula_code, thickness, line_code, quantity, duration, target_days, scenario_tag in extra_long_orders:
        inventory, monthly = build_inventory_fields(quantity, product_code, target_days)
        compat_value = "双拉2线;双拉4线" if product_code != "QDJY" else "双拉2线"
        compat_strategy = "dual-synthetic"
        ws.append([
            order_id,
            f"VAL-MAT-{product_code}-{thickness}",
            f"验证膜_T{thickness}_{product_code}_LONG",
            quantity,
            round(quantity * 0.92, 2),
            0,
            BASE_START.strftime(DATE_FMT),
            BASE_START.strftime(DATE_FMT),
            line_code,
            formula_code,
            thickness,
            product_code,
            duration,
            inventory,
            monthly,
            compat_value,
            f"SYNTHETIC|{scenario_tag}",
            build_validation_tags(target_days, compat_strategy, duration),
            target_days,
            compat_strategy,
            "synthetic-long-order",
        ])
        synthetic_count += 1

    stats["synthetic_rows"] = synthetic_count
    style_sheet(ws)
    return ws.max_row - 1, stats


def create_exception_sheet(target_wb: Workbook) -> int:
    ws = target_wb.create_sheet("停机计划")
    ws.append(["lineId", "lineCode", "startTime", "endTime", "scenarioTag", "note"])
    rows = [
        ("L2", "双拉2线", "2026-03-14 08:00", "2026-03-14 20:00", "HC3-L2-maint-1", "双拉2线首轮整线维护"),
        ("L2", "双拉2线", "2026-03-18 08:00", "2026-03-18 14:00", "HC3-L2-maint-2", "双拉2线半日保养"),
        ("L2", "双拉2线", "2026-03-24 20:00", "2026-03-25 08:00", "HC3-L2-night-stop", "跨夜停机"),
        ("L4", "双拉4线", "2026-03-15 08:00", "2026-03-15 18:00", "HC3-L4-maint-1", "双拉4线白班停机"),
        ("L4", "双拉4线", "2026-03-20 12:00", "2026-03-20 22:00", "HC3-L4-maint-2", "双拉4线延长维护"),
        ("L4", "双拉4线", "2026-03-29 08:00", "2026-03-29 16:00", "HC3-L4-holiday", "节假日窗口"),
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    return len(rows)


def create_filter_sheet(target_wb: Workbook) -> int:
    ws = target_wb.create_sheet("过滤器更换计划")
    ws.append(["lineId", "lineCode", "changeTime", "downtimeMinutes", "scenarioTag", "note"])
    rows = [
        ("L4", "双拉4线", "2026-03-13 18:00", 120, "MC1-L4-window-A", "配合 MC1 排序验证"),
        ("L4", "双拉4线", "2026-03-28 08:00", 90, "MC1-L4-window-B", "第二轮过滤器窗口"),
        ("L2", "双拉2线", "2026-03-16 08:00", 120, "MC1-L2-window-A", "双拉2线过滤器窗口"),
        ("L2", "双拉2线", "2026-03-30 08:00", 60, "MC1-L2-window-B", "双拉2线补充窗口"),
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    return len(rows)


def create_calendar_sheet(target_wb: Workbook) -> int:
    ws = target_wb.create_sheet("工厂日历")
    ws.append(["date", "isWorkingDay", "workStart", "workEnd", "note"])

    holiday_overrides = {
        date(2026, 3, 21): "周六盘点停工",
        date(2026, 4, 4): "清明假期",
    }

    current = BASE_START.date()
    end = current + timedelta(days=34)
    rows = 0
    while current <= end:
        is_working_day = current.weekday() != 6 and current not in holiday_overrides
        note = holiday_overrides.get(current, "常规工作日" if is_working_day else "周日停工")
        ws.append([
            current.isoformat(),
            "Y" if is_working_day else "N",
            "08:00" if is_working_day else "",
            "20:00" if is_working_day else "",
            note,
        ])
        rows += 1
        current += timedelta(days=1)

    style_sheet(ws)
    return rows


def create_coverage_sheet(target_wb: Workbook) -> int:
    ws = target_wb.create_sheet("约束覆盖矩阵")
    ws.append(["constraintId", "当前代码状态", "数据来源", "覆盖方式", "说明"])
    rows = [
        ("HC1", "已启用", "订单_验证版", "单线兼容 + 双线兼容混合", "验证兼容关系是否真正成为集合"),
        ("HC2", "保留未启用", "订单_验证版", "通过库存/发货量反推出 <10 天订单", "可直接计算 inventorySupplyDays"),
        ("HC3", "保留未启用", "停机计划", "两条线分别布置维护窗口", "启用后可验证停机冲突"),
        ("HC4", "保留未启用", "订单_验证版", "保留真实厚度分布并追加极值组合", "便于验证厚度旋转"),
        ("HC5", "保留未启用", "订单_验证版", "大量 >24h 订单 + 两条超长订单", "启用拆分后可验证子任务保序"),
        ("MC1", "保留未启用", "订单_验证版 + 过滤器更换计划", "补齐优先序组合并建立 20 天窗口", "现有 468 行不足以完整覆盖 MC1"),
        ("MC2", "保留未启用", "订单_验证版", "构造 >30 天库存覆盖订单", "可验证高库存不前移"),
        ("SC1", "已启用", "订单_验证版", "保留真实型号/厚度分布", "当前换型主逻辑可直接使用"),
        ("SC2", "保留未启用", "订单_验证版", "构造多个库存覆盖层级", "可验证库存优先级"),
        ("SC3", "保留未启用", "订单_验证版", "保留期望开始时间字段", "计划完工列目前仍未进入模型"),
        ("SC4", "保留未启用", "订单_验证版", "线别作为 preferredLineCode，兼容线单独扩展", "可验证产线偏好"),
        ("SC5", "保留未启用", "订单_验证版", "超长订单作为拆分连续生产候选", "启用拆分后可验证"),
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    return len(rows)


def create_summary_sheet(target_wb: Workbook, stats: Counter[str], order_rows: int, exception_rows: int, filter_rows: int, calendar_rows: int, coverage_rows: int) -> None:
    ws = target_wb.create_sheet("统计摘要")
    ws.append(["metric", "value"])
    rows = [
        ("source_workbook", str(SOURCE_WORKBOOK.relative_to(REPO_ROOT))),
        ("base_order_rows", stats["base_rows"]),
        ("synthetic_order_rows", stats["synthetic_rows"]),
        ("output_order_rows", order_rows),
        ("urgent_rows_lt_10d", stats["urgent_rows"]),
        ("high_stock_rows_gt_30d", stats["high_stock_rows"]),
        ("dual_compatible_rows", stats["dual_compatible_rows"]),
        ("single_compatible_rows", stats["single_compatible_rows"]),
        ("exception_rows", exception_rows),
        ("filter_plan_rows", filter_rows),
        ("calendar_rows", calendar_rows),
        ("coverage_rows", coverage_rows),
        ("generated_at", datetime.now().strftime(DATETIME_FMT)),
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def main() -> None:
    if not SOURCE_WORKBOOK.exists():
        raise FileNotFoundError(f"source workbook not found: {SOURCE_WORKBOOK}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_wb = load_workbook(SOURCE_WORKBOOK, data_only=True)
    source_ws = source_wb[source_wb.sheetnames[0]]

    target_wb = Workbook()
    default_ws = target_wb.active
    target_wb.remove(default_ws)

    order_rows, stats = create_orders_sheet(source_ws, target_wb)
    exception_rows = create_exception_sheet(target_wb)
    filter_rows = create_filter_sheet(target_wb)
    calendar_rows = create_calendar_sheet(target_wb)
    coverage_rows = create_coverage_sheet(target_wb)
    create_summary_sheet(target_wb, stats, order_rows, exception_rows, filter_rows, calendar_rows, coverage_rows)

    target_wb.save(OUTPUT_WORKBOOK)
    print(OUTPUT_WORKBOOK)


if __name__ == "__main__":
    main()
